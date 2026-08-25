import csv

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from app.cleaner.schema import ALL_TABLES


TRACE_COLUMNS = ("fuente", "fecha_extraccion", "fecha_carga")


def add_trace_metadata(
    rows: list[dict],
    fuente: str,
    fecha_extraccion: str,
) -> list[dict]:

    fecha_carga = datetime.now(timezone.utc).isoformat()

    return [
        {
            **row,
            "fuente": fuente,
            "fecha_extraccion": fecha_extraccion,
            "fecha_carga": fecha_carga,
        }
        for row in rows
    ]


def export_table_to_csv(
    rows: list[dict],
    columns: tuple[str, ...],
    output_path: Path,
) -> None:

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as handle:

        writer = csv.DictWriter(handle, fieldnames=list(columns))

        writer.writeheader()

        for row in rows:
            writer.writerow(row)


def export_table_to_sheet(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    columns: tuple[str, ...],
    rows: list[dict],
) -> None:

    sheet = workbook.create_sheet(title=sheet_name[:31])

    sheet.append(list(columns))

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:

        values = [row.get(column) for column in columns]

        excel_row = [
            str(value) if isinstance(value, (list, dict)) else value
            for value in values
        ]

        sheet.append(excel_row)


def export_tables(
    mapped_tables: dict[str, list[dict]],
    output_dir: Path,
    fuente: str,
    fecha_extraccion: str,
) -> dict[str, str]:
    """
    Exporta cada sub-tabla mapeada a su propio CSV en `output_dir`
    (listo para cargarse a PostgreSQL) y además a un único Excel
    consolidado (`datos_limpios.xlsx`, una hoja por sub-tabla) para
    revisión manual. Devuelve {nombre_tabla: ruta_csv}.
    """

    written: dict[str, str] = {}

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for spec in ALL_TABLES:

        rows = mapped_tables.get(spec.name, [])

        if not rows:
            continue

        traced_rows = add_trace_metadata(
            rows,
            fuente=fuente,
            fecha_extraccion=fecha_extraccion,
        )

        columns = tuple(
            column.name for column in spec.columns
        ) + TRACE_COLUMNS

        output_path = output_dir / f"{spec.name}.csv"

        export_table_to_csv(
            traced_rows,
            columns,
            output_path,
        )

        written[spec.name] = str(output_path)

        export_table_to_sheet(workbook, spec.name, columns, traced_rows)

    unmatched = mapped_tables.get("sin_clasificar", [])

    if unmatched:

        output_path = output_dir / "sin_clasificar.csv"

        unmatched_columns = (
            "title", "headers", "source_file", "source_sheet", "row_count",
        )

        export_table_to_csv(
            unmatched,
            unmatched_columns,
            output_path,
        )

        written["sin_clasificar"] = str(output_path)

        export_table_to_sheet(
            workbook, "sin_clasificar", unmatched_columns, unmatched,
        )

    if workbook.sheetnames:

        excel_path = output_dir / "datos_limpios.xlsx"

        workbook.save(excel_path)

        written["excel"] = str(excel_path)

    return written
