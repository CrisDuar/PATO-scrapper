from dataclasses import dataclass, field

import openpyxl

from app.cleaner.text_utils import normalize_text


@dataclass
class DataBlock:
    """
    Representa una tabla individual dentro de una hoja: un título
    opcional, una fila de encabezados y las filas de datos asociadas.
    """

    title: str
    headers: list[str]
    rows: list[list]
    source_sheet: str = ""
    source_file: str = ""
    extra: dict = field(default_factory=dict)


def _row_is_empty(row: tuple) -> bool:

    return all(
        cell is None or normalize_text(cell) == ""
        for cell in row
    )


def _looks_like_header(row: tuple) -> bool:
    """
    Una fila de encabezado tiene casi todas sus celdas no vacías
    y en su mayoría son texto (no números), a diferencia de las
    filas de datos que suelen mezclar texto con números/años.
    """

    values = [
        normalize_text(cell)
        for cell in row
        if normalize_text(cell) != ""
    ]

    if len(values) < 2:
        return False

    non_numeric = sum(
        1
        for value in values
        if not _is_numeric(value)
    )

    return non_numeric >= len(values) - 1


def _is_numeric(value: str) -> bool:

    try:
        float(value.replace(",", "."))
        return True

    except ValueError:
        return False


YEAR_HEADER_TERMS = {"ano", "anio", "year"}


def _is_year_value(value: str) -> bool:
    """
    '2018', '2020*', '2020**' (el DANE marca años estimados/ajustados
    con asteriscos) cuentan como año si sus primeros 4 caracteres son
    dígitos de un año plausible.
    """

    digits = "".join(
        char
        for char in value[:4]
        if char.isdigit()
    )

    return len(digits) == 4 and 1900 <= int(digits) <= 2100


def _looks_like_year_row(row: tuple) -> bool:

    values = [
        normalize_text(cell)
        for cell in row
        if normalize_text(cell) != ""
    ]

    if len(values) < 2:
        return False

    return all(_is_year_value(value) for value in values)


def _unpivot_wide_years(
    headers: list[str],
    year_row: tuple,
    data_rows: list[list],
) -> tuple[list[str], list[list]]:
    """
    Convierte una tabla en formato ancho (una columna por año) a
    formato largo (una fila por año): dado un encabezado terminado
    en 'Año'/'Year' y una fila con los años reales debajo, produce
    encabezados = [*columnas_fijas, 'Anio', 'Valor'] y una fila por
    cada combinación (fila original, columna-año).

    La fila de años suele quedar recortada a la izquierda (sin las
    columnas fijas, que están vacías en esa fila), por lo que sus
    años se alinean con las columnas de datos que vienen *después*
    de las columnas fijas, en el mismo orden relativo.
    """

    fixed_headers = headers[:-1]

    n_fixed = len(fixed_headers)

    years = [
        int("".join(char for char in normalize_text(cell)[:4] if char.isdigit()))
        for cell in year_row
        if _is_year_value(normalize_text(cell))
    ]

    new_headers = [*fixed_headers, "Anio", "Valor"]

    new_rows: list[list] = []

    for row in data_rows:

        fixed_values = list(row[:n_fixed])

        value_columns = row[n_fixed:]

        for year, value in zip(years, value_columns):

            if normalize_text(value) == "":
                continue

            new_rows.append([*fixed_values, year, value])

    return new_headers, new_rows


GROUP_HEADER_TERMS = {
    "dominio",
    "dominios",
    "principales_dominios",
    "pais",
    "paises",
    "region",
    "regiones",
    "departamento",
    "departamentos",
}


def _slug(value: str) -> str:

    import re
    import unicodedata

    text = normalize_text(value).lower()

    text = "".join(
        char
        for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )

    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _looks_like_group_row(row: tuple) -> bool:
    """
    Fila con los nombres de las columnas-grupo (p. ej. Nacional,
    Cabecera, Centros poblados...), típicamente precedida por
    celdas vacías que corresponden a las columnas fijas.
    """

    values = [
        normalize_text(cell)
        for cell in row
        if normalize_text(cell) != ""
    ]

    if len(values) < 2:
        return False

    return all(not _is_numeric(value) for value in values)


def _fix_blank_characteristic_header(headers: list[str]) -> list[str]:
    """
    Corrige un patrón específico del DANE: una columna rotulada
    'Características de la persona' / 'Características del jefe de
    hogar' (que agrupa una etiqueta constante como 'Sexo' en la
    columna) seguida de una columna con encabezado vacío que en
    realidad contiene el valor real (Hombre/Mujer). Renombra esa
    columna vacía a partir del rótulo de la columna anterior, para
    que quede identificable como columna propia (p. ej. 'Sexo').
    """

    fixed = list(headers)

    for index in range(len(fixed) - 1):

        current_key = _slug(fixed[index])

        next_is_blank = normalize_text(fixed[index + 1]) == ""

        if current_key.startswith("caracteristicas_de") and next_is_blank:

            if "jefe" in current_key:
                fixed[index + 1] = "Sexo Jefe Hogar"
            else:
                fixed[index + 1] = "Sexo Persona"

    return fixed


def _unpivot_wide_groups(
    headers: list[str],
    group_row: tuple,
    data_rows: list[list],
) -> tuple[list[str], list[list]]:
    """
    Convierte una tabla matricial (fila = categoría, columna = grupo
    como Dominio/País/Región) a formato largo: dado un encabezado
    terminado en 'Dominio'/'País'/etc. y una fila con los nombres de
    grupo debajo, produce encabezados = [*columnas_fijas,
    <nombre_del_grupo>, 'Valor'] con una fila por cada combinación
    (fila original, columna-grupo).
    """

    fixed_headers = headers[:-1]

    group_column_name = headers[-1]

    n_fixed = len(fixed_headers)

    groups = [
        normalize_text(cell)
        for cell in group_row
        if normalize_text(cell) != ""
    ]

    new_headers = [*fixed_headers, group_column_name, "Valor"]

    new_rows: list[list] = []

    for row in data_rows:

        fixed_values = list(row[:n_fixed])

        value_columns = row[n_fixed:]

        for group_name, value in zip(groups, value_columns):

            if normalize_text(value) == "":
                continue

            new_rows.append([*fixed_values, group_name, value])

    return new_headers, new_rows


def extract_blocks_from_sheet(
    rows: list[tuple],
    sheet_name: str,
    source_file: str,
) -> list[DataBlock]:
    """
    Recorre una hoja fila por fila detectando el patrón:
    [fila(s) vacías] -> [título opcional] -> [encabezado] ->
    [fila de años opcional, si el encabezado termina en 'Año'] ->
    [filas de datos] y devuelve un DataBlock por cada tabla
    encontrada. Tablas en formato ancho (un año por columna) se
    despivotan a formato largo (una fila por año).
    """

    blocks: list[DataBlock] = []

    pending_title = ""

    headers: list[str] | None = None

    pending_year_row: tuple | None = None

    pending_group_row: tuple | None = None

    current_rows: list[list] = []

    def flush():

        if not headers or not current_rows:
            return

        final_headers = _fix_blank_characteristic_header(headers)
        final_rows = current_rows

        last_header_key = _slug(final_headers[-1]) if final_headers else ""

        if pending_year_row is not None and last_header_key in YEAR_HEADER_TERMS:

            final_headers, final_rows = _unpivot_wide_years(
                final_headers,
                pending_year_row,
                current_rows,
            )

        elif pending_group_row is not None and last_header_key in GROUP_HEADER_TERMS:

            final_headers, final_rows = _unpivot_wide_groups(
                final_headers,
                pending_group_row,
                current_rows,
            )

        if not final_rows:
            return

        blocks.append(
            DataBlock(
                title=pending_title,
                headers=final_headers,
                rows=[row[:] for row in final_rows],
                source_sheet=sheet_name,
                source_file=source_file,
            )
        )

    for row in rows:

        trimmed = _squeeze_merged_row(_trim_row(row))

        if _row_is_empty(trimmed):

            flush()

            headers = None
            pending_year_row = None
            pending_group_row = None
            current_rows = []
            continue

        non_empty = [
            normalize_text(cell)
            for cell in trimmed
            if normalize_text(cell) != ""
        ]

        is_title_row = len(non_empty) == 1

        is_new_header_row = (
            headers is not None
            and current_rows
            and _looks_like_header(trimmed)
            and normalize_text(trimmed[0]) != normalize_text(current_rows[-1][0])
        )

        if headers is None or is_title_row or is_new_header_row:

            if is_title_row:
                flush()

                new_title = non_empty[0]

                # Cuando varios títulos-de-una-celda aparecen seguidos
                # sin bloque de datos en medio (p. ej. un título genérico
                # 'Contribuciones...' seguido de uno más específico con
                # el año), se conserva el más largo/informativo en vez
                # de perder el contexto quedándose solo con el último.
                if headers is None and pending_title and not current_rows:
                    pending_title = max(
                        (pending_title, new_title),
                        key=len,
                    )
                else:
                    pending_title = new_title

                headers = None
                pending_year_row = None
                pending_group_row = None
                current_rows = []
                continue

            if _looks_like_header(trimmed):
                flush()
                headers = [
                    normalize_text(cell)
                    for cell in trimmed
                ]
                pending_year_row = None
                pending_group_row = None
                current_rows = []
                continue

            if headers is None:
                continue

        last_header_key = _slug(headers[-1])

        if (
            not current_rows
            and pending_year_row is None
            and last_header_key in YEAR_HEADER_TERMS
            and _looks_like_year_row(trimmed)
        ):
            pending_year_row = trimmed
            continue

        if (
            not current_rows
            and pending_group_row is None
            and last_header_key in GROUP_HEADER_TERMS
            and _looks_like_group_row(trimmed)
        ):
            pending_group_row = trimmed
            continue

        current_rows.append(list(trimmed))

    flush()

    return blocks


def _trim_row(row: tuple) -> tuple:
    """
    Quita columnas vacías al principio/final de la fila (las hojas
    del formato de referencia suelen dejar una columna en blanco
    antes de empezar los datos reales).
    """

    values = list(row)

    while values and normalize_text(values[0]) == "":
        values.pop(0)

    while values and normalize_text(values[-1]) == "":
        values.pop()

    return tuple(values)


def _squeeze_merged_row(row: list) -> tuple:
    """
    Colapsa repeticiones consecutivas idénticas causadas por el
    forward-fill de celdas combinadas (una celda de título fusionada
    horizontalmente A1:H1 se lee como 8 copias del mismo valor; una
    etiqueta de fila fusionada verticalmente como 'Sexo' en A37:B38
    se lee como 2 copias en la misma fila). Sin esto, una fila de
    título de una sola celda deja de detectarse como tal.
    """

    squeezed: list = []

    for value in row:

        if squeezed and normalize_text(value) == normalize_text(squeezed[-1]) and normalize_text(value) != "":
            continue

        squeezed.append(value)

    return tuple(squeezed)


def _read_sheet_rows_with_merges(sheet) -> list[list]:
    """
    Lee una hoja como matriz de valores, propagando el valor de cada
    celda combinada (merged cell) que fusiona *solo filas de datos*
    (dos o más filas, todas ellas ya con datos numéricos en el resto
    de sus columnas) hacia abajo. Esto rellena etiquetas de fila
    fusionadas verticalmente en el cuerpo de la tabla (p. ej. 'Sexo'
    sobre las filas Hombre/Mujer) sin tocar merges que viven en la
    zona de encabezados (donde el forward-fill rompería la detección
    de headers/filas de año, ver block_extractor tests).
    """

    rows = [list(row) for row in sheet.iter_rows(values_only=True)]

    for merged_range in sheet.merged_cells.ranges:

        spans_one_column = merged_range.min_col == merged_range.max_col

        spans_multiple_rows = merged_range.min_row != merged_range.max_row

        if not (spans_one_column and spans_multiple_rows):
            continue

        covered_rows = range(merged_range.min_row, merged_range.max_row + 1)

        all_rows_are_data = all(
            any(
                isinstance(cell, (int, float))
                for cell in rows[row_index - 1]
            )
            for row_index in covered_rows
        )

        if not all_rows_are_data:
            continue

        value = rows[merged_range.min_row - 1][merged_range.min_col - 1]

        for row_index in covered_rows:
            rows[row_index - 1][merged_range.min_col - 1] = value

    _forward_fill_orphan_first_column(rows)

    return rows


def _forward_fill_orphan_first_column(rows: list[list]) -> None:
    """
    Propaga hacia abajo el valor de la primera columna cuando una
    fila de datos la deja en blanco sin que haya un merge real que lo
    explique (el DANE hace esto con frecuencia: la etiqueta de
    'Dominio' solo aparece en la primera fila de cada grupo de
    variables, p. ej. 'Nacional' seguido de filas con la columna de
    Dominio vacía para 'Bajo logro educativo', 'Rezago escolar', …).
    Solo aplica cuando la fila "huérfana" tiene la misma forma que la
    fila anterior (2ª columna con texto, resto numérico) — así no
    contamina filas de encabezado o filas realmente distintas.
    """

    for row_index in range(1, len(rows)):

        row = rows[row_index]
        previous = rows[row_index - 1]

        if not row or not previous:
            continue

        if len(row) < 3 or len(previous) < 3:
            continue

        first_is_blank = normalize_text(row[0]) == ""

        previous_first_has_value = normalize_text(previous[0]) != ""

        second_is_text = (
            normalize_text(row[1]) != ""
            and not isinstance(row[1], (int, float))
        )

        rest_is_numeric = all(
            isinstance(cell, (int, float))
            for cell in row[2:]
            if normalize_text(cell) != ""
        )

        rest_has_values = any(
            normalize_text(cell) != "" for cell in row[2:]
        )

        previous_second_is_text = (
            normalize_text(previous[1]) != ""
            and not isinstance(previous[1], (int, float))
        )

        if (
            first_is_blank
            and previous_first_has_value
            and second_is_text
            and rest_is_numeric
            and rest_has_values
            and previous_second_is_text
        ):
            row[0] = previous[0]


def extract_blocks_from_workbook(path: str) -> list[DataBlock]:

    workbook = openpyxl.load_workbook(
        path,
        data_only=True,
    )

    blocks: list[DataBlock] = []

    for sheet in workbook.worksheets:

        rows = _read_sheet_rows_with_merges(sheet)

        blocks.extend(
            extract_blocks_from_sheet(
                rows,
                sheet_name=sheet.title,
                source_file=path,
            )
        )

    return blocks


def extract_blocks_from_csv(path: str) -> list[DataBlock]:

    import csv

    with open(
        path,
        "r",
        encoding="utf-8",
        errors="replace",
        newline="",
    ) as handle:

        reader = csv.reader(handle)

        rows = [tuple(row) for row in reader]

    return extract_blocks_from_sheet(
        rows,
        sheet_name="csv",
        source_file=path,
    )
